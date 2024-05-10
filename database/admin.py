import random
import time

from django.contrib import admin, messages
from django.contrib.admin.helpers import ACTION_CHECKBOX_NAME
from django.http import HttpResponse, JsonResponse
from django.utils.html import format_html
from database import models
from openpyxl import Workbook
from utils import requests_pro

CONT = "1"


def change_time(timeStamp):
    """
        :param timeStamp: 时间戳
        :return: 时间格式
    """
    timeArray = time.localtime(timeStamp)
    otherStyleTime = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
    minuteTime = time.strftime("%H:%M", timeArray)
    dateTime = time.strftime("%Y-%m-%d",timeArray)
    return otherStyleTime, minuteTime,dateTime


@admin.register(models.PieData)
class PieData(admin.ModelAdmin):

    list_display = ('game_picture','title','tname','view','author','danmaku','like','share','favorite','reply','pub_location','pubdate','href_link','collection')
    list_filter = ['title','tname',]

    list_per_page = 15

    actions = ["export_as_excel",'request_data']

    def game_picture(self, models_obj):
        return format_html('<img src="{}" height="50" width="50">', '{}'.format(models_obj.src))
    game_picture.short_description = '封面'

    # 导出数据
    def export_as_excel(self, request, queryset):
        meta = self.model._meta  # 用于定义文件名, 格式为: app名.模型类名
        field_names = [field.name for field in meta.fields]  # 模型所有字段名
        response = HttpResponse(content_type='application/msexcel')            # 定义响应内容类型
        response['Content-Disposition'] = f'attachment; filename={meta}.xlsx'  # 定义响应数据格式
        wb = Workbook()  # 新建Workbook
        ws = wb.active  # 使用当前活动的Sheet表
        ws.append(field_names)  # 将模型字段名作为标题写入第一行
        for obj in queryset:  # 遍历选择的对象列表
            for field in field_names:
                data = [f'{getattr(obj, field)}' for field in field_names]  # 将模型属性值的文本格式组成列表
            ws.append(data)  # 写入模型属性值
        wb.save(response)  # 将数据存入响应内容
        return response
    export_as_excel.short_description = '导出表格'  # 该动作在admin中的显示文字
    export_as_excel.type = 'warning'
    export_as_excel.icon = 'el-icon-download'

    #  自定义按钮不勾选也可以执行
    def changelist_view(self, request, extra_context=None):
        if 'action' in request.POST and request.POST['action'] == 'request_data':  # 自定义按钮名称
            if not request.POST.getlist(ACTION_CHECKBOX_NAME):
                post = request.POST.copy()
                models.PieData.objects.create(title=CONT,tname=CONT,short_link=CONT,view=CONT,
                                                  author=CONT,src=CONT,danmaku=CONT,like=CONT,share=CONT,
                                                  favorite=CONT,reply=CONT,)
                for u in models.PieData.objects.all():
                    post.update({ACTION_CHECKBOX_NAME: str(u.id)})
                request._set_post(post)
        return super(PieData, self).changelist_view(request, extra_context)

    # 爬取数据
    def request_data(self,request,queryset):
        try:
            for _ in range(1,5):
                session = requests_pro.session_requests_b()
                page = session.get(url='https://api.bilibili.com/x/web-interface/popular?ps=50&pn={}'.format(_))
                page.encoding = "UTF-8"
                for obj in page.json().get('data').get('list'):
                    try:
                        title = obj.get('title')
                        tname = obj.get('tname')  # 标签
                        short_link = obj.get('short_link_v2')  # 连接
                        view = obj.get('stat').get('view')  # 播放次数
                        author = obj.get('owner').get('name')  # 作者
                        src = obj.get('pic')  # 封面
                        danmaku = obj.get('stat').get('danmaku')  # 弹幕数
                        like = obj.get('stat').get('like')  # 点赞数
                        share = obj.get('stat').get('share')  # 分享数
                        favorite = obj.get('stat').get('favorite')  # 收藏数
                        pubdate = obj.get('pubdate')  # 发布时间
                        pub_location = obj.get('pub_location')
                        reply = obj.get('stat').get('reply')  # 评论数
                        print("视频: {}  添加成功".format(title))
                        if models.PieData.objects.filter(title=title).exists() == False:
                            models.PieData.objects.create(title=title,tname=tname,short_link=short_link,view=view,
                                                          author=author,src=src,danmaku=danmaku,like=like,share=share,pub_location=pub_location,
                                                          favorite=favorite,pubdate=change_time(pubdate)[2],reply=reply)
                        if models.History.objects.filter(title=title).exists() == False:
                            models.History.objects.create(title=title, tname=tname, short_link=short_link, view=view,
                                                          author=author, src=src, danmaku=danmaku, like=like,
                                                          share=share,pub_location=pub_location,
                                                          favorite=favorite, pubdate=change_time(pubdate)[2],
                                                          reply=reply)
                    except Exception as e:
                        print(e)

            models.PieData.objects.filter(title="1").delete()
            return HttpResponse("爬取成功，请刷新页面")

        except Exception as e:
            return HttpResponse("爬取失败，失败原因:{},请联系管理员处理(QQ995405033)".format(e),)

    request_data.short_description = '爬取数据'
    request_data.type = 'success'
    request_data.icon = 'el-icon-s-promotion'






@admin.register(models.BarData)
class BarData(admin.ModelAdmin):

    list_display = ('game_picture','title','tname','view','author','danmaku','like','share','favorite','reply','pub_location','pubdate','href_link','collection')
    list_filter = ['title','tname',]

    list_per_page = 20
    actions = ["export_as_excel",'request_data']

    #
    def game_picture(self, models_obj):
        return format_html('<img src="{}" height="50" width="50">', '{}'.format(models_obj.src))
    game_picture.short_description = '图片'

    # 导出数据
    def export_as_excel(self, request, queryset):
        meta = self.model._meta  # 用于定义文件名, 格式为: app名.模型类名
        field_names = [field.name for field in meta.fields]  # 模型所有字段名
        response = HttpResponse(content_type='application/msexcel')  # 定义响应内容类型
        response['Content-Disposition'] = f'attachment; filename={meta}.xlsx'  # 定义响应数据格式
        wb = Workbook()  # 新建Workbook
        ws = wb.active  # 使用当前活动的Sheet表
        ws.append(field_names)  # 将模型字段名作为标题写入第一行
        for obj in queryset:  # 遍历选择的对象列表
            for field in field_names:
                data = [f'{getattr(obj, field)}' for field in field_names]  # 将模型属性值的文本格式组成列表
            ws.append(data)  # 写入模型属性值
        wb.save(response)  # 将数据存入响应内容
        return response
    export_as_excel.short_description = '导出表格'  # 该动作在admin中的显示文字
    export_as_excel.type = 'warning'
    export_as_excel.icon = 'el-icon-download'

    #  自定义按钮不勾选也可以执行
    def changelist_view(self, request, extra_context=None):
        if 'action' in request.POST and request.POST['action'] == 'request_data':  # 自定义按钮名称
            if not request.POST.getlist(ACTION_CHECKBOX_NAME):
                post = request.POST.copy()
                models.BarData.objects.create(title=CONT, tname=CONT, short_link=CONT, view=CONT,
                                              author=CONT, src=CONT, danmaku=CONT, like=CONT, share=CONT,
                                              favorite=CONT, reply=CONT)
                for u in models.BarData.objects.all():
                    post.update({ACTION_CHECKBOX_NAME: str(u.id)})
                request._set_post(post)
        return super(BarData, self).changelist_view(request, extra_context)


    # 爬取数据
    def request_data(self,request,queryset):
        try:
            for _ in range(190,201):
                session = requests_pro.session_requests_b()
                page = session.get(url='https://api.bilibili.com/x/web-interface/popular/series/one?number={}'.format(_))
                page.encoding = "UTF-8"
                for obj in page.json().get('data').get('list'):
                    try:
                        title = obj.get('title')
                        tname = obj.get('tname')  # 标签
                        short_link = obj.get('short_link_v2')  # 连接
                        view = obj.get('stat').get('view')  # 播放次数
                        author = obj.get('owner').get('name')  # 作者
                        src = obj.get('pic')  # 封面
                        danmaku = obj.get('stat').get('danmaku')  # 弹幕数
                        like = obj.get('stat').get('like')  # 点赞数
                        share = obj.get('stat').get('share')  # 分享数
                        favorite = obj.get('stat').get('favorite')  # 收藏数
                        pubdate = obj.get('pubdate')  # 发布时间
                        pub_location = obj.get('pub_location')
                        reply = obj.get('stat').get('reply')  # 评论数
                        print("视频: {}  添加成功".format(title))
                        if models.BarData.objects.filter(title=title).exists() == False:
                            models.BarData.objects.create(title=title,tname=tname,short_link=short_link,view=view,
                                                          author=author,src=src,danmaku=danmaku,like=like,share=share,pub_location=pub_location,
                                                          favorite=favorite,pubdate=change_time(pubdate)[2],reply=reply)
                        if models.History.objects.filter(title=title).exists() == False:
                            models.History.objects.create(title=title, tname=tname, short_link=short_link, view=view,
                                                          author=author, src=src, danmaku=danmaku, like=like,
                                                          share=share,pub_location=pub_location,
                                                          favorite=favorite, pubdate=change_time(pubdate)[2],
                                                          reply=reply)
                    except Exception as e:
                        print(e)

            models.BarData.objects.filter(title="1").delete()
            return HttpResponse("爬取成功，请刷新页面")

        except Exception as e:
            return HttpResponse("爬取失败，失败原因:{},请联系管理员处理(QQ995405033)".format(e),)

    request_data.short_description = '爬取数据'
    request_data.type = 'success'
    request_data.icon = 'el-icon-s-promotion'








@admin.register(models.LintData)
class LintData(admin.ModelAdmin):

    list_display = ('game_picture','title','tname','view','author','danmaku','like','share','favorite','reply','pub_location','pubdate','href_link','collection')
    list_filter = ['title','tname',]

    list_per_page = 20
    actions = ["export_as_excel",'request_data']

    # 图片
    def game_picture(self, models_obj):
        return format_html('<img src="{}" height="50" width="50">', '{}'.format(models_obj.src))
    game_picture.short_description = '图片'

    # 导出数据
    def export_as_excel(self, request, queryset):
        meta = self.model._meta  # 用于定义文件名, 格式为: app名.模型类名
        field_names = [field.name for field in meta.fields]  # 模型所有字段名
        response = HttpResponse(content_type='application/msexcel')  # 定义响应内容类型
        response['Content-Disposition'] = f'attachment; filename={meta}.xlsx'  # 定义响应数据格式
        wb = Workbook()  # 新建Workbook
        ws = wb.active  # 使用当前活动的Sheet表
        ws.append(field_names)  # 将模型字段名作为标题写入第一行
        for obj in queryset:  # 遍历选择的对象列表
            for field in field_names:
                data = [f'{getattr(obj, field)}' for field in field_names]  # 将模型属性值的文本格式组成列表
            ws.append(data)  # 写入模型属性值
        wb.save(response)  # 将数据存入响应内容
        return response
    export_as_excel.short_description = '导出表格'  # 该动作在admin中的显示文字
    export_as_excel.type = 'warning'
    export_as_excel.icon = 'el-icon-download'

    #  自定义按钮不勾选也可以执行
    def changelist_view(self, request, extra_context=None):
        if 'action' in request.POST and request.POST['action'] == 'request_data':  # 自定义按钮名称
            if not request.POST.getlist(ACTION_CHECKBOX_NAME):
                post = request.POST.copy()
                models.LintData.objects.create(title=CONT, tname=CONT, short_link=CONT, view=CONT,
                                              author=CONT, src=CONT, danmaku=CONT, like=CONT, share=CONT,
                                              favorite=CONT, reply=CONT)
                for u in models.LintData.objects.all():
                    post.update({ACTION_CHECKBOX_NAME: str(u.id)})
                request._set_post(post)
        return super(LintData, self).changelist_view(request, extra_context)


    # 爬取数据
    def request_data(self,request,queryset):
        try:
            for _ in range(1,5):
                session = requests_pro.session_requests_b()
                page = session.get(url='https://api.bilibili.com/x/web-interface/popular/precious?page_size=100&page={}'.format(_))
                page.encoding = "UTF-8"
                for obj in page.json().get('data').get('list'):
                    try:
                        title = obj.get('title')
                        tname = obj.get('tname')  # 标签
                        short_link = obj.get('short_link_v2')  # 连接
                        view = obj.get('stat').get('view')  # 播放次数
                        author = obj.get('owner').get('name')  # 作者
                        src = obj.get('pic')  # 封面
                        danmaku = obj.get('stat').get('danmaku')  # 弹幕数
                        like = obj.get('stat').get('like')  # 点赞数
                        share = obj.get('stat').get('share')  # 分享数
                        favorite = obj.get('stat').get('favorite')  # 收藏数
                        pubdate = obj.get('pubdate')  # 发布时间
                        pub_location = obj.get('pub_location')
                        reply = obj.get('stat').get('reply')  # 评论数
                        print("视频: {}  添加成功".format(title))
                        if models.LintData.objects.filter(title=title).exists() == False:
                            models.LintData.objects.create(title=title,tname=tname,short_link=short_link,view=view,
                                                          author=author,src=src,danmaku=danmaku,like=like,share=share,pub_location=pub_location,
                                                          favorite=favorite,pubdate=change_time(pubdate)[2],reply=reply)
                        if models.History.objects.filter(title=title).exists() == False:
                            models.History.objects.create(title=title, tname=tname, short_link=short_link, view=view,
                                                          author=author, src=src, danmaku=danmaku, like=like,
                                                          share=share,pub_location=pub_location,
                                                          favorite=favorite, pubdate=change_time(pubdate)[2],
                                                          reply=reply)
                    except Exception as e:
                        print(e)
            models.LintData.objects.filter(title="1").delete()
            return HttpResponse("爬取成功，请刷新页面")

        except Exception as e:
            return HttpResponse("爬取失败，失败原因:{},请联系管理员处理(QQ995405033)".format(e),)

    request_data.short_description = '爬取数据'
    request_data.type = 'success'
    request_data.icon = 'el-icon-s-promotion'




@admin.register(models.BarsData)
class BarsData(admin.ModelAdmin):

    list_display = ('game_picture','title','tname','view','author','danmaku','like','share','favorite','reply','pub_location','pubdate','href_link','collection')
    list_filter = ['title','tname',]

    list_per_page = 20
    actions = ["export_as_excel",'request_data']

    def game_picture(self, models_obj):
        return format_html('<img src="{}" height="50" width="50">', '{}'.format(models_obj.src))
    game_picture.short_description = '封面'

    # 导出数据
    def export_as_excel(self, request, queryset):
        meta = self.model._meta  # 用于定义文件名, 格式为: app名.模型类名
        field_names = [field.name for field in meta.fields]  # 模型所有字段名
        response = HttpResponse(content_type='application/msexcel')  # 定义响应内容类型
        response['Content-Disposition'] = f'attachment; filename={meta}.xlsx'  # 定义响应数据格式
        wb = Workbook()  # 新建Workbook
        ws = wb.active  # 使用当前活动的Sheet表
        ws.append(field_names)  # 将模型字段名作为标题写入第一行
        for obj in queryset:  # 遍历选择的对象列表
            for field in field_names:
                data = [f'{getattr(obj, field)}' for field in field_names]  # 将模型属性值的文本格式组成列表
            ws.append(data)  # 写入模型属性值
        wb.save(response)  # 将数据存入响应内容
        return response
    export_as_excel.short_description = '导出表格'  # 该动作在admin中的显示文字
    export_as_excel.type = 'warning'
    export_as_excel.icon = 'el-icon-download'

    #  自定义按钮不勾选也可以执行
    def changelist_view(self, request, extra_context=None):
        if 'action' in request.POST and request.POST['action'] == 'request_data':  # 自定义按钮名称
            if not request.POST.getlist(ACTION_CHECKBOX_NAME):
                post = request.POST.copy()
                models.BarsData.objects.create(title=CONT, tname=CONT, short_link=CONT, view=CONT,
                                              author=CONT, src=CONT, danmaku=CONT, like=CONT, share=CONT,
                                              favorite=CONT, reply=CONT)
                for u in models.BarsData.objects.all():
                    post.update({ACTION_CHECKBOX_NAME: str(u.id)})
                request._set_post(post)
        return super(BarsData, self).changelist_view(request, extra_context)


    # 爬取数据
    def request_data(self,request,queryset):
        try:
            for _ in range(0,2):
                session = requests_pro.session_requests_b()
                page = session.get(url='https://api.bilibili.com/x/web-interface/ranking/v2?rid={}&type=all'.format(_))
                page.encoding = "UTF-8"
                for obj in page.json().get('data').get('list'):
                    try:
                        title = obj.get('title')
                        tname = obj.get('tname')  # 标签
                        short_link = obj.get('short_link_v2')  # 连接
                        view = obj.get('stat').get('view')  # 播放次数
                        author = obj.get('owner').get('name')  # 作者
                        src = obj.get('pic')  # 封面
                        danmaku = obj.get('stat').get('danmaku')  # 弹幕数
                        like = obj.get('stat').get('like')  # 点赞数
                        share = obj.get('stat').get('share')  # 分享数
                        favorite = obj.get('stat').get('favorite')  # 收藏数
                        pubdate = obj.get('pubdate')  # 发布时间
                        pub_location = obj.get('pub_location')
                        reply = obj.get('stat').get('reply')  # 评论数
                        print("视频: {}  添加成功".format(title))
                        if models.BarsData.objects.filter(title=title).exists() == False:
                            models.BarsData.objects.create(title=title,tname=tname,short_link=short_link,view=view,
                                                          author=author,src=src,danmaku=danmaku,like=like,share=share,pub_location=pub_location,
                                                          favorite=favorite,pubdate=change_time(pubdate)[2],reply=reply)
                        if models.History.objects.filter(title=title).exists() == False:
                            models.History.objects.create(title=title,tname=tname,short_link=short_link,view=view,
                                                          author=author,src=src,danmaku=danmaku,like=like,share=share,pub_location=pub_location,
                                                          favorite=favorite,pubdate=change_time(pubdate)[2],reply=reply)
                    except Exception as e:
                        print(e)
            models.BarsData.objects.filter(title="1").delete()
            return HttpResponse("爬取成功，请刷新页面")
        except Exception as e:
            return HttpResponse("爬取失败，失败原因:{},请联系管理员处理(QQ995405033)".format(e),)

    request_data.short_description = '爬取数据'
    request_data.type = 'success'
    request_data.icon = 'el-icon-s-promotion'





@admin.register(models.History)
class History(admin.ModelAdmin):

    list_display = ('game_picture','title','tname','view','author','danmaku','like','share','favorite','reply','pub_location','pubdate','href_link','collection')
    list_filter = ['title','tname',]

    list_per_page = 20
    actions = ["export_as_excel",'request_data']

    def game_picture(self, models_obj):
        return format_html('<img src="{}" height="50" width="50">', '{}'.format(models_obj.src))
    game_picture.short_description = '封面'

    # 导出数据
    def export_as_excel(self, request, queryset):
        meta = self.model._meta  # 用于定义文件名, 格式为: app名.模型类名
        field_names = [field.name for field in meta.fields]  # 模型所有字段名
        response = HttpResponse(content_type='application/msexcel')  # 定义响应内容类型
        response['Content-Disposition'] = f'attachment; filename={meta}.xlsx'  # 定义响应数据格式
        wb = Workbook()  # 新建Workbook
        ws = wb.active  # 使用当前活动的Sheet表
        ws.append(field_names)  # 将模型字段名作为标题写入第一行
        for obj in queryset:  # 遍历选择的对象列表
            for field in field_names:
                data = [f'{getattr(obj, field)}' for field in field_names]  # 将模型属性值的文本格式组成列表
            ws.append(data)  # 写入模型属性值
        wb.save(response)  # 将数据存入响应内容
        return response
    export_as_excel.short_description = '导出表格'  # 该动作在admin中的显示文字
    export_as_excel.type = 'warning'
    export_as_excel.icon = 'el-icon-download'





@admin.register(admin.models.LogEntry)
class LogEntryAdmin(admin.ModelAdmin):

    list_display = ['action_time', 'user', 'content_type', '__str__']
    list_display_links = ['action_time']
    list_filter = ['action_time', 'content_type', 'user']
    list_per_page = 15
    readonly_fields = ['action_time', 'user', 'content_type','object_id', 'object_repr', 'action_flag', 'change_message']



class Colect(admin.ModelAdmin):   # 这里需要修改

    list_display = ("game_picture","co_title","co_down_link","href_link")
    list_filter = ["co_title","co_down_link"]
    list_per_page = 20
    actions = ["export_as_excel"]

    # 图片
    def game_picture(self, models_obj):
        return format_html('<img src="{}" height="50" width="50">', '{}'.format(models_obj.co_down_img))
    game_picture.short_description = '图片'

    def export_as_excel(self, request, queryset):
        meta = self.model._meta  # 用于定义文件名, 格式为: app名.模型类名
        field_names = [field.name for field in meta.fields]  # 模型所有字段名
        response = HttpResponse(content_type='application/msexcel')  # 定义响应内容类型
        response['Content-Disposition'] = f'attachment; filename={meta}.xlsx'  # 定义响应数据格式
        wb = Workbook()  # 新建Workbook
        ws = wb.active  # 使用当前活动的Sheet表
        ws.append(field_names)  # 将模型字段名作为标题写入第一行
        for obj in queryset:  # 遍历选择的对象列表

            for field in field_names:
                data = [f'{getattr(obj, field)}' for field in field_names]  # 将模型属性值的文本格式组成列表
            ws.append(data)
        wb.save(response)
        return response

    export_as_excel.short_description = '导出Excel'  # 该动作在admin中的显示文字

admin.site.register(models.Colect, Colect)  # 这里需要修改





admin.site.site_title="抖音短视频数据分析系统"
admin.site.site_header="抖音短视频数据分析系统"
admin.site.index_title="抖音短视频数据分析系统"